import csv
import io
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "transaction_date": [
        "completed date",
        "started date",
        "date",
        "created date",
        "settled date",
        "timestamp",
    ],
    "amount": [
        "amount",
        "gross amount",
        "payment amount",
        "amount (gbp)",
        "amount (eur)",
        "amount (usd)",
    ],
    "currency": ["currency", "base currency", "account currency"],
    "description": ["description", "details", "merchant", "notes", "comment"],
    "counterparty": ["counterparty", "beneficiary", "payee", "from", "to", "card"],
    "transaction_type": ["type", "transaction type", "operation type"],
    "status": ["status", "state"],
    "external_id": ["id", "transaction id", "operation id"],
    "raw_reference": ["reference", "payment reference", "bank reference"],
}


def _normalize_header(value: str) -> str:
    normalized = "".join(ch.lower() if ch.isalnum() else " " for ch in value)
    return " ".join(normalized.split())


def _select_key(normalized_headers: dict[str, str], aliases: list[str]) -> str | None:
    for alias in aliases:
        key = _normalize_header(alias)
        if key in normalized_headers:
            return normalized_headers[key]
    return None


def parse_statement_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        rows = _parse_csv(content)
    elif lower_name.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Unsupported file type. Please upload a CSV or XLSX export.")

    records = [_extract_record(row) for row in rows]
    return [record for record in records if record is not None]


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("Could not decode CSV file.")

    reader = csv.DictReader(io.StringIO(decoded))
    return list(reader)


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    clean_headers = []
    for idx, value in enumerate(headers):
        header = str(value).strip() if value is not None else ""
        clean_headers.append(header or f"column_{idx + 1}")

    rows: list[dict[str, Any]] = []
    for row_values in rows_iter:
        row_dict: dict[str, Any] = {}
        for idx, value in enumerate(row_values):
            if idx < len(clean_headers):
                row_dict[clean_headers[idx]] = value
        rows.append(row_dict)
    return rows


def _extract_record(row: dict[str, Any]) -> dict[str, Any] | None:
    normalized_headers = {_normalize_header(key): key for key in row.keys()}

    field_values: dict[str, Any] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        key = _select_key(normalized_headers, aliases)
        field_values[field_name] = row.get(key) if key else None

    transaction_date = _parse_date(field_values.get("transaction_date"))
    amount = _parse_amount(field_values.get("amount"))

    if transaction_date is None or amount is None:
        return None

    return {
        "transaction_date": transaction_date,
        "amount": amount,
        "currency": _coerce_str(field_values.get("currency")),
        "description": _coerce_str(field_values.get("description")),
        "counterparty": _coerce_str(field_values.get("counterparty")),
        "transaction_type": _coerce_str(field_values.get("transaction_type")),
        "status": _coerce_str(field_values.get("status")),
        "external_id": _coerce_str(field_values.get("external_id")),
        "raw_reference": _coerce_str(field_values.get("raw_reference")),
    }


def _coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized).date()
    except ValueError:
        pass

    known_formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d %b %Y",
        "%d %B %Y",
    ]
    for fmt in known_formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    cleaned = "".join(ch for ch in text if ch.isdigit() or ch in ".,-")

    if cleaned.count(",") > 0 and cleaned.count(".") == 0:
        cleaned = cleaned.replace(",", ".")
    elif cleaned.count(",") > 0 and cleaned.count(".") > 0:
        cleaned = cleaned.replace(",", "")

    try:
        amount = float(cleaned)
    except ValueError:
        return None

    if negative:
        amount *= -1
    return amount
